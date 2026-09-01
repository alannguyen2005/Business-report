"""
check_data.py
-------------
Cổng kiểm tra trước khi chạy phân tích. Trả lời đúng một câu hỏi:
"Có nên chạy pipeline phân tích lúc này không?"

Data kinh doanh không tự kéo về — người dùng đặt file vào đường dẫn
khai báo ở `data.file` trong config.yaml.

Kiểm tra:
  1. File data tồn tại và đọc được
  2. File có đúng các cột khai báo trong config `schema`
  3. Data đã đổi so với lần phân tích gần nhất (so fingerprint)
  4. Data không quá cũ (nếu `schedule.max_data_age_days` > 0)

Exit code:
  0  → nên chạy (in ra RUN)
  10 → bỏ qua, không phải lỗi (in ra SKIP: lý do)
  1  → lỗi thật, cần người xử lý (in ra ERROR: lý do)

Dùng:
  python check_data.py            # kiểm tra bình thường
  python check_data.py --force    # bỏ qua check fingerprint, luôn RUN
  python check_data.py --json     # in kết quả dạng JSON
  python check_data.py --record   # ghi state sau khi phân tích xong
"""

import argparse
import csv
import hashlib
import json
import sys
from datetime import date, datetime
from pathlib import Path

import yaml

# Console Windows mặc định cp1252 — ép UTF-8 để in được tiếng Việt
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

ROOT = Path(__file__).resolve().parent
CONFIG_FILE = ROOT / "config.yaml"
STATE_FILE = ROOT / "logs" / "last_analyzed.json"

EXIT_RUN, EXIT_ERROR, EXIT_SKIP = 0, 1, 10


# ─── HELPERS ─────────────────────────────────────────────────────────────────

def load_config() -> dict:
    return yaml.safe_load(CONFIG_FILE.read_text(encoding="utf-8"))


def fingerprint(path: Path) -> str:
    """SHA-256 của nội dung file — đổi 1 byte là đổi fingerprint."""
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()


def load_state() -> dict:
    if not STATE_FILE.exists():
        return {}
    try:
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except (json.JSONDecodeError, OSError):
        return {}


def read_header(path: Path) -> list[str]:
    if path.suffix.lower() in {".xlsx", ".xls"}:
        try:
            from openpyxl import load_workbook
        except ImportError:
            return []          # không kiểm được thì bỏ qua, không chặn pipeline
        ws = load_workbook(path, read_only=True).active
        return [str(c.value) for c in next(ws.iter_rows(max_row=1)) if c.value]
    with open(path, newline="", encoding="utf-8-sig") as f:
        return next(csv.reader(f), [])


def latest_date_in_csv(path: Path, col: str, fmt: str) -> date | None:
    """Ngày mới nhất trong cột thời gian. Trả None nếu không parse được."""
    if path.suffix.lower() in {".xlsx", ".xls"}:
        return None            # bỏ qua với Excel — không đáng để mở cả file
    latest = None
    with open(path, newline="", encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            raw = (row.get(col) or "").strip()
            if not raw:
                continue
            parsed = None
            for candidate in (fmt, "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    parsed = datetime.strptime(raw.split(" ")[0], candidate).date()
                    break
                except ValueError:
                    continue
            if parsed and (latest is None or parsed > latest):
                latest = parsed
    return latest


def required_columns(schema: dict) -> list[str]:
    """Các cột bắt buộc phải có: date, revenue, cộng các chiều đã khai báo."""
    cols = [schema.get("date"), schema.get("revenue")]
    cols += schema.get("dimensions") or []
    return [c for c in cols if c]


# ─── MAIN ────────────────────────────────────────────────────────────────────

def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--force", action="store_true",
                        help="Bỏ qua check fingerprint, luôn chạy")
    parser.add_argument("--json", action="store_true",
                        help="In kết quả dạng JSON")
    parser.add_argument("--record", action="store_true",
                        help="Ghi lại fingerprint sau khi phân tích xong (không kiểm tra gì)")
    args = parser.parse_args()

    result: dict = {"decision": None, "reason": None}

    def emit(decision: str, reason: str, code: int, **extra) -> int:
        result.update(decision=decision, reason=reason, **extra)
        if args.json:
            print(json.dumps(result, ensure_ascii=False, indent=2))
        else:
            print(f"{decision}: {reason}")
        return code

    if not CONFIG_FILE.exists():
        return emit("ERROR", f"Không tìm thấy {CONFIG_FILE.name}", EXIT_ERROR)

    config = load_config()
    data_path = ROOT / config["data"]["file"]
    schema = config.get("schema", {})

    # Chế độ ghi state — gọi sau khi pipeline phân tích chạy xong
    if args.record:
        if not data_path.exists():
            return emit("ERROR", f"Không tìm thấy data file: {config['data']['file']}",
                        EXIT_ERROR)
        STATE_FILE.parent.mkdir(parents=True, exist_ok=True)
        state = {
            "fingerprint": fingerprint(data_path),
            "analyzed_at": date.today().isoformat(),
            "data_file": config["data"]["file"],
            "stem": config["data"]["stem"],
        }
        latest_recorded = None
        if schema.get("date"):
            latest_recorded = latest_date_in_csv(
                data_path, schema["date"], schema.get("date_format", "%Y-%m-%d"))
        if latest_recorded:
            state["data_latest_date"] = latest_recorded.isoformat()
        STATE_FILE.write_text(json.dumps(state, ensure_ascii=False, indent=2),
                              encoding="utf-8")
        return emit("RECORDED",
                    f"Đã ghi state vào {STATE_FILE.relative_to(ROOT)} "
                    f"(fingerprint {state['fingerprint'][:12]}...)",
                    EXIT_RUN, **state)

    # 1. File tồn tại
    if not data_path.exists():
        return emit("ERROR",
                    f"Không tìm thấy data file: {config['data']['file']}. "
                    "Kiểm tra lại `data.file` trong config.yaml.",
                    EXIT_ERROR)

    result["data_file"] = str(config["data"]["file"])

    # 2. Schema khớp
    header = read_header(data_path)
    if header:
        missing = [c for c in required_columns(schema) if c not in header]
        if missing:
            return emit("ERROR",
                        f"Data thiếu cột khai báo trong config `schema`: {', '.join(missing)}. "
                        f"Cột có trong file: {', '.join(header[:12])}...",
                        EXIT_ERROR, missing_columns=missing)

    # 3. Độ tươi của data (chỉ khi được bật)
    max_age = int(config.get("schedule", {}).get("max_data_age_days", 0) or 0)
    latest = None
    if schema.get("date"):
        latest = latest_date_in_csv(data_path, schema["date"],
                                    schema.get("date_format", "%Y-%m-%d"))
    if latest:
        result["data_latest_date"] = latest.isoformat()
        age = (date.today() - latest).days
        result["data_age_days"] = age
        if max_age > 0 and age > max_age:
            return emit("SKIP",
                        f"Data cũ {age} ngày (mới nhất {latest}), quá ngưỡng "
                        f"{max_age} ngày. Nguồn data có thể chưa cập nhật.",
                        EXIT_SKIP)

    # 4. Data đã đổi chưa
    current_fp = fingerprint(data_path)
    result["fingerprint"] = current_fp
    state = load_state()

    if args.force:
        return emit("RUN", "Chạy cưỡng bức (--force), bỏ qua check fingerprint",
                    EXIT_RUN)

    if state.get("fingerprint") == current_fp:
        return emit("SKIP",
                    f"Data không đổi kể từ lần phân tích {state.get('analyzed_at', '?')}. "
                    "Thêm data mới hoặc chạy với --force để phân tích lại.",
                    EXIT_SKIP, last_analyzed=state.get("analyzed_at"))

    return emit("RUN",
                "Data mới hoặc đã thay đổi — nên chạy phân tích"
                if state else "Chưa từng phân tích data này",
                EXIT_RUN, last_analyzed=state.get("analyzed_at"))


if __name__ == "__main__":
    sys.exit(main())
